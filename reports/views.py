from django.shortcuts import render
from django.contrib.admin.views.decorators import staff_member_required
from django.http import HttpResponse, JsonResponse
from django.db.models import Q, Count, Avg, Sum, F
from django.db.models.functions import TruncMonth
from warehouse.models import Asset, AssetReferral, CartridgeCharge, Supplier, AssetCategory
from personnel.models import Personnel, Branch, Department
from tickets.models import Ticket, TicketCategory
from projects.models import Project
import datetime


@staff_member_required
def dashboard(request):
    """داشبورد مدیریتی"""
    today = datetime.date.today()
    month_start = today.replace(day=1)

    # آمار کلی
    context = {
        'total_assets': Asset.objects.count(),
        'total_personnel': Personnel.objects.filter(is_active=True).count(),
        'total_branches': Branch.objects.filter(is_active=True).count(),
        'total_tickets': Ticket.objects.count(),
        'open_tickets': Ticket.objects.filter(status='open').count(),
        'in_progress_tickets': Ticket.objects.filter(status='in_progress').count(),
        'resolved_tickets': Ticket.objects.filter(status='resolved').count(),
        'closed_tickets': Ticket.objects.filter(status='closed').count(),
        'total_projects': Project.objects.count(),
        'total_suppliers': Supplier.objects.filter(is_active=True).count(),

        # آمار تجهیزات
        'fixed_assets': Asset.objects.filter(asset_type='fixed').count(),
        'consumable_assets': Asset.objects.filter(asset_type='consumable').count(),
        'under_repair': Asset.objects.filter(status='under_repair').count(),
        'under_charge': Asset.objects.filter(status='under_charge').count(),

        # آمار ارجاعات
        'pending_referrals': AssetReferral.objects.filter(status='pending').count(),
        'active_referrals': AssetReferral.objects.filter(status__in=['sent', 'in_progress']).count(),

        # آمار شارژ کارتریج
        'pending_charges': CartridgeCharge.objects.filter(status__in=['sent', 'charging']).count(),

        # نمودار تیکت‌ها بر اساس وضعیت (برای Chart.js)
        'ticket_status_chart': {
            'labels': ['باز', 'در حال انجام', 'حل شده', 'بسته شده'],
            'data': [
                Ticket.objects.filter(status='open').count(),
                Ticket.objects.filter(status='in_progress').count(),
                Ticket.objects.filter(status='resolved').count(),
                Ticket.objects.filter(status='closed').count(),
            ],
        },

        # نمودار تیکت‌ها بر اساس شعبه
        'ticket_branch_chart': _get_ticket_by_branch(),

        # نمودار تیکت‌ها بر اساس دسته‌بندی
        'ticket_category_chart': _get_ticket_by_category(),

        # نمودار ارجاعات تجهیزات
        'referral_type_chart': _get_referral_by_type(),

        # تیکت‌های اخیر
        'recent_tickets': Ticket.objects.select_related('requester', 'branch', 'target_branch').order_by('-created_at')[:10],

        # ارجاعات اخیر
        'recent_referrals': AssetReferral.objects.select_related('asset', 'supplier').order_by('-created_at')[:10],
    }
    return render(request, 'reports/dashboard.html', context)


def _get_ticket_by_branch():
    """تیکت‌ها بر اساس شعبه"""
    branches = Ticket.objects.values('branch__name').annotate(count=Count('id')).order_by('-count')
    labels = [b['branch__name'] or 'نامشخص' for b in branches]
    data = [b['count'] for b in branches]
    return {'labels': labels, 'data': data}


def _get_ticket_by_category():
    """تیکت‌ها بر اساس دسته‌بندی"""
    categories = Ticket.objects.values('category__title').annotate(count=Count('id')).order_by('-count')[:10]
    labels = [c['category__title'] or 'نامشخص' for c in categories]
    data = [c['count'] for c in categories]
    return {'labels': labels, 'data': data}


def _get_referral_by_type():
    """ارجاعات بر اساس نوع"""
    types = AssetReferral.objects.values('referral_type').annotate(count=Count('id'))
    type_labels = {'repair': 'تعمیر', 'charge': 'شارژ', 'upgrade': 'ارتقا', 'scrap': 'اوراق', 'transfer': 'انتقال', 'other': 'سایر'}
    labels = [type_labels.get(t['referral_type'], t['referral_type']) for t in types]
    data = [t['count'] for t in types]
    return {'labels': labels, 'data': data}


@staff_member_required
def asset_report(request):
    """گزارش تجهیزات"""
    assets = Asset.objects.select_related('supplier', 'warehouse', 'branch', 'asset_category').all()

    # فیلترها
    search = request.GET.get('search', '')
    branch_id = request.GET.get('branch', '')
    asset_type = request.GET.get('asset_type', '')
    status = request.GET.get('status', '')
    supplier_id = request.GET.get('supplier', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    if search:
        assets = assets.filter(Q(name__icontains=search) | Q(code__icontains=search) | Q(part_number__icontains=search))
    if branch_id:
        assets = assets.filter(branch_id=branch_id)
    if asset_type:
        assets = assets.filter(asset_type=asset_type)
    if status:
        assets = assets.filter(status=status)
    if supplier_id:
        assets = assets.filter(supplier_id=supplier_id)
    if date_from:
        assets = assets.filter(purchase_date__gte=date_from)
    if date_to:
        assets = assets.filter(purchase_date__lte=date_to)

    context = {
        'assets': assets,
        'branches': Branch.objects.filter(is_active=True),
        'suppliers': Supplier.objects.filter(is_active=True),
        'filters': {
            'search': search, 'branch': branch_id, 'asset_type': asset_type,
            'status': status, 'supplier': supplier_id, 'date_from': date_from, 'date_to': date_to,
        },
        'total_price': assets.aggregate(total=Sum('price'))['total'] or 0,
        'report_title': 'گزارش تجهیزات',
    }
    return render(request, 'reports/asset_report.html', context)


@staff_member_required
def consumable_report(request):
    """گزارش کالاهای مصرفی"""
    consumables = Asset.objects.filter(asset_type='consumable').select_related('supplier', 'branch')

    search = request.GET.get('search', '')
    branch_id = request.GET.get('branch', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    if search:
        consumables = consumables.filter(Q(name__icontains=search) | Q(code__icontains=search))
    if branch_id:
        consumables = consumables.filter(branch_id=branch_id)
    if date_from:
        consumables = consumables.filter(usage_start_date__gte=date_from)
    if date_to:
        consumables = consumables.filter(usage_start_date__lte=date_to)

    context = {
        'consumables': consumables,
        'branches': Branch.objects.filter(is_active=True),
        'filters': {'search': search, 'branch': branch_id, 'date_from': date_from, 'date_to': date_to},
        'report_title': 'گزارش کالاهای مصرفی',
    }
    return render(request, 'reports/consumable_report.html', context)


@staff_member_required
def cartridge_report(request):
    """گزارش شارژ کارتریج"""
    charges = CartridgeCharge.objects.select_related('asset', 'supplier', 'branch', 'sent_by', 'received_by').all()

    search = request.GET.get('search', '')
    supplier_id = request.GET.get('supplier', '')
    branch_id = request.GET.get('branch', '')
    status = request.GET.get('status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    if search:
        charges = charges.filter(Q(asset__name__icontains=search) | Q(code__icontains=search))
    if supplier_id:
        charges = charges.filter(supplier_id=supplier_id)
    if branch_id:
        charges = charges.filter(branch_id=branch_id)
    if status:
        charges = charges.filter(status=status)
    if date_from:
        charges = charges.filter(send_date__gte=date_from)
    if date_to:
        charges = charges.filter(send_date__lte=date_to)

    # آمار
    total_cost = charges.aggregate(total=Sum('cost'))['total'] or 0
    avg_quality = charges.filter(quality_rating__isnull=False).aggregate(avg=Avg('quality_rating'))['avg'] or 0
    avg_speed = charges.filter(speed_rating__isnull=False).aggregate(avg=Avg('speed_rating'))['avg'] or 0

    context = {
        'charges': charges,
        'suppliers': Supplier.objects.filter(is_active=True),
        'branches': Branch.objects.filter(is_active=True),
        'filters': {'search': search, 'supplier': supplier_id, 'branch': branch_id, 'status': status, 'date_from': date_from, 'date_to': date_to},
        'total_cost': total_cost,
        'avg_quality': round(avg_quality, 1),
        'avg_speed': round(avg_speed, 1),
        'report_title': 'گزارش شارژ کارتریج',
    }
    return render(request, 'reports/cartridge_report.html', context)


@staff_member_required
def referral_report(request):
    """گزارش ارجاعات تجهیزات"""
    referrals = AssetReferral.objects.select_related('asset', 'supplier', 'sent_by', 'received_by').all()

    search = request.GET.get('search', '')
    referral_type = request.GET.get('referral_type', '')
    status = request.GET.get('status', '')
    supplier_id = request.GET.get('supplier', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    if search:
        referrals = referrals.filter(Q(asset__name__icontains=search) | Q(code__icontains=search) | Q(asset__code__icontains=search))
    if referral_type:
        referrals = referrals.filter(referral_type=referral_type)
    if status:
        referrals = referrals.filter(status=status)
    if supplier_id:
        referrals = referrals.filter(supplier_id=supplier_id)
    if date_from:
        referrals = referrals.filter(send_date__gte=date_from)
    if date_to:
        referrals = referrals.filter(send_date__lte=date_to)

    context = {
        'referrals': referrals,
        'suppliers': Supplier.objects.filter(is_active=True),
        'REFERRAL_TYPES': AssetReferral.REFERRAL_TYPES,
        'filters': {'search': search, 'referral_type': referral_type, 'status': status, 'supplier': supplier_id, 'date_from': date_from, 'date_to': date_to},
        'total_cost': referrals.aggregate(total=Sum('cost'))['total'] or 0,
        'report_title': 'گزارش ارجاعات تجهیزات',
    }
    return render(request, 'reports/referral_report.html', context)


@staff_member_required
def ticket_report(request):
    """گزارش تیکت‌ها"""
    tickets = Ticket.objects.select_related('requester', 'assigned_to', 'category', 'project', 'branch', 'target_branch').all()

    search = request.GET.get('search', '')
    status = request.GET.get('status', '')
    priority = request.GET.get('priority', '')
    category_id = request.GET.get('category', '')
    branch_id = request.GET.get('branch', '')
    target_branch_id = request.GET.get('target_branch', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    if search:
        tickets = tickets.filter(Q(title__icontains=search) | Q(code__icontains=search) | Q(requester__full_name__icontains=search))
    if status:
        tickets = tickets.filter(status=status)
    if priority:
        tickets = tickets.filter(priority=priority)
    if category_id:
        tickets = tickets.filter(category_id=category_id)
    if branch_id:
        tickets = tickets.filter(branch_id=branch_id)
    if target_branch_id:
        tickets = tickets.filter(target_branch_id=target_branch_id)
    if date_from:
        tickets = tickets.filter(created_at__date__gte=date_from)
    if date_to:
        tickets = tickets.filter(created_at__date__lte=date_to)

    # آمار
    total = tickets.count()
    resolved = tickets.filter(status__in=['resolved', 'closed']).count()

    context = {
        'tickets': tickets,
        'categories': TicketCategory.objects.all(),
        'branches': Branch.objects.filter(is_active=True),
        'filters': {
            'search': search, 'status': status, 'priority': priority,
            'category': category_id, 'branch': branch_id, 'target_branch': target_branch_id,
            'date_from': date_from, 'date_to': date_to,
        },
        'total': total,
        'resolved': resolved,
        'resolution_rate': round(resolved / total * 100, 1) if total > 0 else 0,
        'report_title': 'گزارش تیکت‌ها',
    }
    return render(request, 'reports/ticket_report.html', context)


@staff_member_required
def supplier_report(request):
    """گزارش شرکت‌ها و تأمین‌کنندگان"""
    suppliers = Supplier.objects.all()

    search = request.GET.get('search', '')
    if search:
        suppliers = suppliers.filter(Q(name__icontains=search) | Q(contact_person__icontains=search))

    # آمار هر شرکت
    supplier_stats = []
    for s in suppliers:
        asset_count = Asset.objects.filter(supplier=s).count()
        referral_count = AssetReferral.objects.filter(supplier=s).count()
        charge_count = CartridgeCharge.objects.filter(supplier=s).count()
        avg_rating = AssetReferral.objects.filter(supplier=s, quality_rating__isnull=False).aggregate(avg=Avg('quality_rating'))['avg']
        total_cost = AssetReferral.objects.filter(supplier=s).aggregate(total=Sum('cost'))['total'] or 0
        total_cost += CartridgeCharge.objects.filter(supplier=s).aggregate(total=Sum('cost'))['total'] or 0

        supplier_stats.append({
            'supplier': s,
            'asset_count': asset_count,
            'referral_count': referral_count,
            'charge_count': charge_count,
            'avg_rating': round(avg_rating, 1) if avg_rating else 0,
            'total_cost': total_cost,
        })

    context = {
        'supplier_stats': supplier_stats,
        'filters': {'search': search},
        'report_title': 'گزارش شرکت‌ها و تأمین‌کنندگان',
    }
    return render(request, 'reports/supplier_report.html', context)


def export_excel(request, report_type):
    """خروجی اکسل"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        return HttpResponse("کتابخانه openpyxl نصب نیست. با پشتیبان تماس بگیرید.", content_type='text/plain; charset=utf-8')

    wb = openpyxl.Workbook()
    ws = wb.active

    # استایل هدر
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4B2D8E", end_color="4B2D8E", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    if report_type == 'assets':
        ws.title = "تجهیزات"
        headers = ['کد', 'نام', 'نوع', 'وضعیت', 'شعبه', 'تأمین‌کننده', 'قیمت', 'تاریخ خرید']
        ws.append(headers)

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        assets = Asset.objects.select_related('supplier', 'branch')
        for a in assets:
            ws.append([
                a.code, a.name, a.get_asset_type_display(), a.get_status_display(),
                a.branch.name if a.branch else '-',
                a.supplier.name if a.supplier else '-',
                float(a.price) if a.price else 0,
                str(a.purchase_date) if a.purchase_date else '-',
            ])

    elif report_type == 'tickets':
        ws.title = "تیکت‌ها"
        headers = ['کد', 'عنوان', 'وضعیت', 'اولویت', 'شعبه', 'واحد IT', 'درخواست‌کننده', 'کارشناس', 'زمان پاسخ', 'تاریخ']
        ws.append(headers)

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        tickets = Ticket.objects.select_related('requester', 'assigned_to', 'branch', 'target_branch', 'category')
        for t in tickets:
            response_time = '-'
            if t.resolved_at and t.created_at:
                diff = t.resolved_at - t.created_at
                hours = diff.total_seconds() / 3600
                if hours < 1:
                    response_time = f"{int(diff.total_seconds() / 60)} دقیقه"
                elif hours < 24:
                    response_time = f"{int(hours)} ساعت"
                else:
                    response_time = f"{int(hours / 24)} روز"

            ws.append([
                t.code, t.title, t.get_status_display(), t.get_priority_display(),
                t.branch.name if t.branch else '-',
                t.target_branch.name if t.target_branch else '-',
                t.requester.full_name if t.requester else '-',
                t.assigned_to.full_name if t.assigned_to else '-',
                response_time,
                str(t.created_at.strftime('%Y-%m-%d %H:%M')) if t.created_at else '-',
            ])

    elif report_type == 'cartridges':
        ws.title = "شارژ کارتریج"
        headers = ['کد', 'کارتریج', 'شرکت شارژ', 'شعبه', 'تاریخ ارسال', 'تاریخ بازگشت', 'هزینه', 'وضعیت', 'امتیاز کیفیت', 'امتیاز سرعت']
        ws.append(headers)

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        charges = CartridgeCharge.objects.select_related('asset', 'supplier', 'branch')
        for c in charges:
            ws.append([
                c.code, c.asset.name, c.supplier.name,
                c.branch.name if c.branch else '-',
                str(c.send_date), str(c.return_date) if c.return_date else '-',
                float(c.cost) if c.cost else 0,
                c.get_status_display(),
                c.quality_rating or '-', c.speed_rating or '-',
            ])

    elif report_type == 'referrals':
        ws.title = "ارجاعات"
        headers = ['کد', 'تجهیز', 'نوع', 'وضعیت', 'شرکت', 'تاریخ ارسال', 'تاریخ بازگشت', 'هزینه', 'امتیاز']
        ws.append(headers)

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        referrals = AssetReferral.objects.select_related('asset', 'supplier')
        for r in referrals:
            ws.append([
                r.code, r.asset.name, r.get_referral_type_display(), r.get_status_display(),
                r.supplier.name if r.supplier else '-',
                str(r.send_date) if r.send_date else '-',
                str(r.return_date) if r.return_date else '-',
                float(r.cost) if r.cost else 0,
                r.quality_rating or '-',
            ])

    # عرض ستون‌ها
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 4, 40)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{report_type}_report.xlsx"'

    wb.save(response)
    return response
